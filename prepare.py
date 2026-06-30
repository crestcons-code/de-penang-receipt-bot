# prepare.py -- Step 1: Parse bank statement and prepare review file
#
# Usage:
#   python prepare.py "507013883446_20260619 (1).csv"
#
# Output: review_YYYYMMDD_HHMMSS.xlsx on your Desktop
# Edit that file, then run:
#   python main.py review_YYYYMMDD_HHMMSS.xlsx

import sys
import os
import pandas as pd
from datetime import datetime
from parse_maybank import load_statement
from donation_mapping import map_to_gl, DONATION_MAP, get_department
from autocount_api import AutocountClient


def prepare(statement_file: str):
    print("\n=== Step 1: Prepare Review File ===")
    print(f"Statement : {statement_file}\n")

    # Parse bank statement
    transactions = load_statement(statement_file)
    print(f"Found {len(transactions)} incoming payment(s).\n")

    client = AutocountClient()

    # Check existing OR records in Autocount for duplicate detection
    from_date = transactions["date"].min().strftime("%Y-%m-%d")
    to_date   = transactions["date"].max().strftime("%Y-%m-%d")
    print(f"Checking Autocount for existing OR records ({from_date} to {to_date})...")
    posted = client.get_posted_receipts(from_date, to_date)
    # Build a lookup set: (date, amount) — amount is most reliable unique key per day
    posted_keys = set()
    for p in posted:
        posted_keys.add((p["date"], p["amount"]))
    print(f"Found {len(posted)} existing OR record(s) in Autocount for this period.\n")

    # Pre-assign OR numbers based on last number in Autocount
    print("Checking last OR number in Autocount...")
    first_date = transactions["date"].iloc[0].strftime("%Y-%m-%d")
    yy = first_date[2:4]
    mm = first_date[5:7]
    prefix = f"OR-{yy}{mm}"
    last_or = client.get_last_or_number(prefix=prefix)
    if last_or and last_or.startswith(prefix):
        next_seq = int(last_or[len(prefix):]) + 1
    else:
        next_seq = 1
    print(f"Last OR: {last_or or 'none'}  ->  next will start at {prefix}{next_seq:03d}\n")

    # Build review rows — skip already-posted transactions
    rows = []
    skipped = []
    seq = next_seq
    for _, txn in transactions.iterrows():
        txn_date = txn["date"].strftime("%Y-%m-%d")
        t_yy, t_mm = txn_date[2:4], txn_date[5:7]
        amount = round(float(txn["credit"]), 2)
        key = (txn_date, amount)

        if key in posted_keys:
            skipped.append(f"  SKIP (already in Autocount): {txn['donor_name']}  RM{amount:.2f}  {txn_date}")
            continue

        or_no = f"OR-{t_yy}{t_mm}{seq:03d}"
        gl_code, gl_name, short_desc = map_to_gl(txn["gl_text"])
        rows.append({
            "OR Number":                  or_no,
            "Post (YES/NO)":              "YES",
            "Date":                       txn_date,
            "Donor Name":                 txn["donor_name"],
            "GL Account Code":            gl_code,
            "GL Description":             gl_name,
            "Description (in Autocount)": short_desc,
            "Department":                 get_department(gl_code),
            "Amount (RM)":                txn["credit"],
        })
        seq += 1

    if skipped:
        print("Skipped (already recorded in Autocount):")
        for s in skipped:
            print(s)
        print()
    print(f"Transactions to review: {len(rows)}  |  Already in Autocount: {len(skipped)}\n")

    df = pd.DataFrame(rows)

    # Save to Desktop
    desktop  = os.path.join(os.path.expanduser("~"), "Desktop")
    out_path = os.path.join(desktop, f"review_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Donations")

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        ws = writer.sheets["Donations"]

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col) + 4
            ws.column_dimensions[col[0].column_letter].width = min(max_len, 50)

        # Find the Amount column index (1-based)
        headers = [cell.value for cell in ws[1]]
        amt_col = headers.index("Amount (RM)") + 1  # 1-based

        # Total row
        total_row = len(df) + 2  # +1 header +1 for next row
        total = df["Amount (RM)"].sum()

        # Write "TOTAL" label in the Donor Name column
        donor_col = headers.index("Donor Name") + 1
        label_cell = ws.cell(row=total_row, column=donor_col, value="TOTAL")
        label_cell.font = Font(bold=True)
        label_cell.alignment = Alignment(horizontal="right")

        # Write total amount
        total_cell = ws.cell(row=total_row, column=amt_col, value=total)
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0.00'
        total_cell.alignment = Alignment(horizontal="right")

        # Top border to separate total row
        thin = Side(style="thin")
        for c in range(1, len(headers) + 1):
            ws.cell(row=total_row, column=c).border = Border(top=thin)

        # Highlight total row
        yellow = PatternFill("solid", fgColor="FFF2CC")
        for c in range(1, len(headers) + 1):
            ws.cell(row=total_row, column=c).fill = yellow

    print("Review file saved to Desktop:")
    print(f"  {out_path}\n")
    print("Instructions:")
    print("  1. Open the file in Excel")
    print("  2. Check each row — correct GL Account Code if needed")
    print("  3. Set 'Post (YES/NO)' to NO for any row you want to skip")
    print("  4. Save the file")
    print(f"  5. Run:  python main.py \"{out_path}\"")
    print()

    # Print available GL codes for reference
    print("=== Available Donation GL Codes ===")
    seen = set()
    for gl_code, short_desc, _ in DONATION_MAP:
        if gl_code not in seen:
            print(f"  {gl_code}  {short_desc}")
            seen.add(gl_code)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python prepare.py <maybank_statement.csv>")
        sys.exit(1)
    prepare(sys.argv[1])
